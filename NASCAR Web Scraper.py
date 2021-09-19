import os, sys, bs4, openpyxl, re, requests, xlsxwriter, pandas as pd, lxml.html as lh, time
from openpyxl import load_workbook, Workbook
from datetime import date
from os.path import exists
    
# Initial definition
scrapeurl = 'http://www.driveraverages.com/nascar/race.php?sked_id='
date = date.today()
blankcheck = 0
book = 0
intcols = [1,2,6,7,8,9,10,11,13]

# Input season
while True:
    try:
        season = int(input("Enter year: ")) # The user types the year of the season.
    except ValueError:
        print("Syntax error, try again.") # It doesn't matter if you type Kyle Busch or Kyle Larson; if it's not an integer, it's a syntax error.
        continue
    if season < 1949:
        print("I don't think that was a very eventful season. Try again.") # 1949 was the first year of NASCAR, so the year has to be 1949 or after.
        continue
    elif season > date.year:
        print("Sorry Biff, you'll have to wait like everyone else.") # If the year entered is in the future, even if the races have already been given titles, the program isn't going to scrape data that isn't there.
        continue
    else:
        break
folderexists = os.path.exists(str(season))
if folderexists == False:
    os.mkdir(str(season))
txtpath = str(season) + '/data.txt'    
txtexists = exists(txtpath)
if txtexists == False:
    print("Creating a file to track web scraping progress . . .")
    makethetxt = open(txtpath,'w+')
    makethetxt.write("0\n0\n0")
    makethetxt.close()
with open(str(season) + '/data.txt','r+') as txt:
    readfile = txt.read()
completed = int(readfile.split('\n', 2)[0]) # Line 1: Completed; Line 2: Races in season; Line 3: Cleaned? [0],[1],[2] respectively
racecount = int(readfile.split('\n', 2)[1])
cleancheck = int(readfile.split('\n', 2)[2])
txt.close()

# Count races
if racecount == 0: # Figure out how many races are in the season.
    print("Determining number of races in the season . . .")
    getcount = requests.get(scrapeurl + str(season) + "001")
    getcount2 = bs4.BeautifulSoup(getcount.content, 'lxml')
    getcount3 = getcount2.find_all('table',{"class": "sortable tabledata-nascar"})
    getcount4 = pd.read_html(str(getcount3)) # Creates a dataframe from the table
    if date.year == season:
        lastline = getcount4[0].values[-1] # If getting the current year, the table should interpret the last line.
        blankcheck = 1 # If the season being scraped is the current year, it's possible that the season is not finished yet. This accounts for that.
    else:
        lastline = getcount4[0].values[-2] # If getting a different year, the table should interpret the next to last line.
    racecount = int(lastline[0])
    with open(str(season) + '/data.txt','w') as txt:
        txt.write(str(completed) + '\n' + str(racecount) + '\n' + str(cleancheck))
        txt.close()
    exit
else:
    exit
print(str(completed) + " races have been scraped from this season.")
print("There are " + str(racecount) + " races in the " + str(season) + " season.")
if cleancheck == 0:
    print("The data has not been cleaned.")
else:
    print("The data has been cleaned.")

if completed == 0:
    print("Commencing data collection . . .")
elif completed > 0 and completed < racecount:
    print("Continuing data collection from Race #" + str(completed + 1) + " . . .")

# Scrape the page for race data and write it to a sheet.
while completed in range(racecount): # Loop for number of races in the season.
      print("Race #" + str(completed+1) + " of " + str(racecount))
      nascar = requests.get(scrapeurl + str(season) + str(completed+1).zfill(3), timeout=None) # Makes the URL to scrape the scrapeurl plus a number with 3 digits.
      pagecontents = bs4.BeautifulSoup(nascar.text, 'html.parser') # Uses BS4 to parse the text of the 'nascar' variable.
      racenameget = pagecontents.select('.td-bold') # Grabs only the td-bold class from the parsed HTML.
      compiledregex = re.compile(r'<span class="td-bold">.+?<br/>') # Define regex: <span class="td-bold">, any text of any length, and a line break.
      ln = compiledregex.search(str(racenameget)) # Searches through 'racenameget' using the conditions of the regex from the last line.
      rn = str(ln.group()) # Converts ln to a string.
      rn1 = rn[22:] # Removes the first 22 characters, which reduces this to the race name plus a line break.
      sheetname = rn1.strip("<br/>") # Strips the line break, leaving the race name itself as the sheet name.
      sheetname = sheetname.replace("/", " ") # Some races have slashes, which worksheet titles don't care for.
      sheetname = sheetname.replace("'", "")
      print(sheetname) # For debug purposes: Print the race name
      soup = bs4.BeautifulSoup(nascar.content, 'lxml') # Parses the page again, but with lxml, to make getting the table easier.
      data = soup.find_all('table',{"class": "sortable tabledata-nascar table-large"}) # Finds tables, but narrows them down to the only one with the sortable tabledata-nascar table-large class.
      df = pd.read_html(str(data)) # Creates a dataframe from the table
      if blankcheck == 1: # If the season being scraped is the current year, we will have to make sure this race has actually happened yet.
          isitblank = df[0].values[0] # Gets the first value of the second row. Is it a 1? Then the race was run.
          if isitblank[0] != str(1): # Checks to see if the value is not a 1, AKA this race hasn't happened yet.
              print("Season incomplete. Stopping here . . .") # If this race didn't run, all race data has been scraped at this point. Move on to cleanup.
              break
      print(df[0]) # For debug purposes: Print the dataframe contents; it's a single element list (contained in brackets), so the [0] pulls that one element.
      wbpath = r"" + str(season) + "/NASCAR_" + str(season) + ".xlsx" # Sets 'wbpath' to the location of the workbook I want to modify
      file_exists = exists(wbpath) # Is there a file at that path?
      if file_exists == False: # No?
          wb = Workbook() # Makes a workbook.
          wb.save(wbpath) # Saves the workbook to wbpath.
      book = load_workbook(wbpath) # Loads the workbook from the specified path, 'wbpath'
      sheetcount = len(book.sheetnames)
      writer = pd.ExcelWriter(wbpath, engine = 'openpyxl') # Specifies the openpyxl engine and points the Excel writer at the workbook.
      writer.book = book # Sets the workbook to the book variable.
      df[0].to_excel(writer, sheet_name=sheetname[0:30], index=False) # Writes the first (only) element in df to a worksheet with the first 31 characters of the 'sheetname' variable.
      worksheet = book[sheetname[0:30]] # Edit the worksheet
      if season < 2005: # If the season was not under the playoffs format...
          worksheet['M1'] = 'Rating' # Change cell M1 value from 'Unnamed: 12' to 'Rating'
      completed += 1    
      writer.save() # Saves the workbook.
      writer.close() # Closes the editor between edits so killing the process halfway through will hypothetically be possible.
      with open(str(season) + '/data.txt','w') as txt:
        txt.write(str(completed) + '\n' + str(racecount) + '\n' + str(cleancheck))
        txt.close()
      time.sleep(2) # This website doesn't always respond to requests; the sleep timer is to refrain from overwhelming the site.

# Clean data
if completed == racecount and cleancheck == 0:
    if book == 0:
        wbpath = r"" + str(season) + "/NASCAR_" + str(season) + ".xlsx" # Sets 'wbpath' to the location of the workbook I want to modify
        book = load_workbook(wbpath)
    print("All race data has been scraped.")
    print("Making final preparations . . .")
    print("Removing borders . . .")
    nostyle = openpyxl.styles.Side(border_style=None) # Define "nostyle" variable as no border.
    no_border = openpyxl.styles.borders.Border(left=nostyle, right=nostyle, top=nostyle, bottom=nostyle) # Remove borders from all cells.
    print("Converting strings to numbers (this may take a while) . . .")
    for sheet in book: # Go through all the sheets.
        r = 2
        c = 1
        sheet.delete_rows(sheet.max_row, 1) # Delete the last row; it's the same as the first row anyway.
        row_count = sheet.max_row
        while r in range(row_count+1) and c in range(14):
            celledit = sheet.cell(row=r,column=c)
            if c in intcols and celledit.data_type != 'n' and celledit.data_type != 'inlineStr':
                if c == 13 and celledit.value != '':
                    celledit.value = float(celledit.value)
                elif c != 13 and celledit.value != '':
                    celledit.value = int(celledit.value)
            r += 1
            if r == row_count + 1:
                r = 2
                c += 1
        for row in sheet:
            for cell in row:
                cell.border = no_border # Remove borders from cells.
    print("Removing useless sheets . . .")
    if 'Sheet' in book.sheetnames:
        getridof = book['Sheet'] # When file_exists detects no file, it creates one. When it does so, it comes with this sheet by default. We wait until openpyxl is done making new worksheets to go in and delete this one.
        book.remove(getridof) # This deletes the sheet.
    print("/" + wbpath + " is now ready!\n")
    book.save(wbpath) # This saves the workbook.
    book.close() # This closes the workbook.
    cleancheck = 1
    with open(str(season) + '/data.txt','w') as txt:
        txt.write(str(completed) + '\n' + str(racecount) + '\n' + str(cleancheck))
        txt.close()
time.sleep(1)
print("Exiting program . . .")
exit()
