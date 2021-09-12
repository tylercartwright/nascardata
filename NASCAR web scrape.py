# Tyler Cartwright - NASCAR web scraper

import bs4, openpyxl, re, requests, xlsxwriter, pandas as pd, lxml.html as lh, time
from openpyxl import load_workbook, Workbook
from datetime import date
from os.path import exists

scrapeurl = 'https://www.driveraverages.com/nascar/race.php?sked_id='
date = date.today()
x = 0 # Initializes x for a while loop that occurs later. 'While' used instead of 'for' so that x can be manipulated in the middle of iteration.
intcols = [1,2,6,7,8,9,10,11,13]
blankcheck = 0 # Initializes this value for eventual use in scraping. Only becomes 1 if this script is being run for the current year and an incomplete season is a possibility.
skiprowdelete = 0 # Initializes this value for eventual use in cleanup. Only becomes 1 if this script is being run on a season that has already been scraped in full.

# Part 1: Input the season to scrape.
while True:
    try:
        season = int(input("Enter year: ")) # The user types the year of the season.
    except ValueError:
        print("Syntax error, try again.") # It doesn't matter if you type Kyle Busch or Kyle Larson; if it's not an integer, it's a syntax error.
        continue
    if season < 1949:
        print("I don't think that was a very eventful season. Try again.") # 1949 was the first year of NASCAR, so the year has to be 1949 or after.
        continue
    elif season > date.year:
        print("Sorry Biff, you'll have to wait like everyone else.") # If the year entered is in the future, even if the races have already been given titles, the program isn't going to scrape data that isn't there.
        continue
    else:
        break

# Part 2: Check how many races are in the season.
getcount = requests.get(scrapeurl + str(season) + "001")
getcount2 = bs4.BeautifulSoup(getcount.content, 'lxml')
getcount3 = getcount2.find_all('table',{"class": "sortable tabledata-nascar"})
getcount4 = pd.read_html(str(getcount3)) # Creates a dataframe from the table
if date.year == season:
    lastline = getcount4[0].values[-1] # If getting the current year, the table should interpret the last line.
    blankcheck = 1 # If the season being scraped is the current year, it's possible that the season is not finished yet. This accounts for that.
else:
    lastline = getcount4[0].values[-2] # If getting a different year, the table should interpret the next to last line.
racecount = int(lastline[0]) # How many races are in the season? The lastline value uses the end of the bottom table to tell us that.
print("Getting data on " + str(racecount) + " races . . . \n")

# Part 3: Scrape the page for race data and write it to a sheet.
while x in range(racecount): # Loop for number of races in the season.
      x += 1
      print("Race #" + str(x) + " of " + str(racecount))
      nascar = requests.get(scrapeurl + str(season) + str(x).zfill(3), timeout=None) # Makes the URL to scrape the scrapeurl plus a number with 3 digits.
      pagecontents = bs4.BeautifulSoup(nascar.text, 'html.parser') # Uses BS4 to parse the text of the 'nascar' variable.
      racenameget = pagecontents.select('.td-bold') # Grabs only the td-bold class from the parsed HTML.
      compiledregex = re.compile(r'<span class="td-bold">.+?<br/>') # Define regex: <span class="td-bold">, any text of any length, and a line break.
      ln = compiledregex.search(str(racenameget)) # Searches through 'racenameget' using the conditions of the regex from the last line.
      rn = str(ln.group()) # Converts ln to a string.
      rn1 = rn[22:] # Removes the first 22 characters, which reduces this to the race name plus a line break.
      sheetname = rn1.strip("<br/>") # Strips the line break, leaving the race name itself as the sheet name.
      sheetname = sheetname.replace("/", " ") # Some races have slashes, which worksheet titles don't care for.
      print(sheetname) # For debug purposes: Print the race name
      soup = bs4.BeautifulSoup(nascar.content, 'lxml') # Parses the page again, but with lxml, to make getting the table easier.
      data = soup.find_all('table',{"class": "sortable tabledata-nascar table-large"}) # Finds tables, but narrows them down to the only one with the sortable tabledata-nascar table-large class.
      df = pd.read_html(str(data)) # Creates a dataframe from the table
      if blankcheck == 1: # If the season being scraped is the current year, we will have to make sure this race has actually happened yet.
          isitblank = df[0].values[0] # Gets the first value of the second row. Is it a 1? Then the race was run.
          if isitblank[0] != str(1): # Checks to see if the value is not a 1, AKA this race hasn't happened yet.
              print("Season incomplete. Stopping here . . .") # If this race didn't run, all race data has been scraped at this point. Move on to cleanup.
              break
      print(df[0]) # For debug purposes: Print the dataframe contents; it's a single element list (contained in brackets), so the [0] pulls that one element.
      wbpath = r"NASCAR_" + str(season) + ".xlsx" # Sets 'wbpath' to the location of the workbook I want to modify
      file_exists = exists(wbpath) # Is there a file at that path?
      if file_exists == False: # No?
          wb = Workbook() # Makes a workbook.
          wb.save(wbpath) # Saves the workbook to wbpath.
      book = load_workbook(wbpath) # Loads the workbook from the specified path, 'wbpath'
      sheetcount = len(book.sheetnames)
      writer = pd.ExcelWriter(wbpath, engine = 'openpyxl') # Specifies the openpyxl engine and points the Excel writer at the workbook.
      writer.book = book # Sets the workbook to the book variable.
      if sheetname[0:30] in book.sheetnames: # If this sheet already exists...
          if x == racecount: # If every race in the season has been scraped already...
              print('This season has already been scraped. Cleaning data . . .') # Move on to cleaning data.
              skiprowdelete = 1 # We don't want to delete the last row at this point; if every race has already been scraped, in theory the cleanup has already run. We don't want to delete important data.
              break
          else:
              print('This race has already been recorded. Skipping ahead . . .') # Skip it.
              x = sheetcount+1 # This is what we used a while loop for. We can jump ahead to the first race we did not scrape.
              print("Starting from race #" + str(x+1)) # x+1 because x will increase at the beginning of the loop.
      else:
          df[0].to_excel(writer, sheet_name=sheetname[0:30], index=False) # Writes the first (only) element in df to a worksheet with the first 31 characters of the 'sheetname' variable.
          worksheet = book[sheetname[0:30]] # Edit the worksheet
          if season < 2005: # If the season was not under the playoffs format...
              worksheet['M1'] = 'Rating' # Change cell M1 value from 'Unnamed: 12' to 'Rating'
          
      writer.save() # Saves the workbook.
      writer.close() # Closes the editor between edits so killing the process halfway through will hypothetically be possible.
      time.sleep(5) # This website doesn't always respond to requests; the sleep timer is to refrain from overwhelming the site.

# Part 4: Clean the data.
print("Making final preparations . . .")
print("Removing borders . . .")
nostyle = openpyxl.styles.Side(border_style=None) # Define "nostyle" variable as no border.
no_border = openpyxl.styles.borders.Border(left=nostyle, right=nostyle, top=nostyle, bottom=nostyle) # Remove borders from all cells.
print("Converting strings to numbers (this may take a while) . . .")
for sheet in book: # Go through all the sheets.
    r = 2
    c = 1
    if skiprowdelete != 1: # If the entire season was not previously scraped...
        sheet.delete_rows(sheet.max_row, 1) # Delete the last row; it's the same as the first row anyway.
    row_count = sheet.max_row
    while r in range(row_count+1) and c in range(14):
        celledit = sheet.cell(row=r,column=c)
        if c in intcols and celledit.data_type != 'n' and celledit.data_type != 'inlineStr':
            if c == 13:
                celledit.value = float(celledit.value)
            else:
                celledit.value = int(celledit.value)
        r += 1
        if r == row_count + 1:
            r = 2
            c += 1
    for row in sheet:
        for cell in row:
            cell.border = no_border # Remove borders from cells.
print("Removing useless sheets . . .")
if 'Sheet' in book.sheetnames:
    getridof = book['Sheet'] # When file_exists detects no file, it creates one. When it does so, it comes with this sheet by default. We wait until openpyxl is done making new worksheets to go in and delete this one.
    book.remove(getridof) # This deletes the sheet.
print(wbpath + " is now ready!")
book.save(wbpath) # This saves the workbook.
book.close() # This closes the workbook.
